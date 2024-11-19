# Kaden Billin
# UWYO COSC 1010
# 11-19-2024
# HW 5
# Sources, people worked with, help given to:
# your
# comments
# here

import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill
import string

wb = op.load_workbook("template.xlsx") # open the workbook to scrape
sheet = wb.active # get the active sheet from it
cells = tuple(sheet['A1':'CV100']) # get your collection of cells in their rows
color_dict = {} # create a dictionary to hold the colors and coordinates
for row in cells:
    for cell in row:
        # Check if the cell has a fill
        if cell.fill and cell.fill.start_color:
            # Extract the fill color
            color = cell.fill.start_color.rgb

            # Get the coordinate
            coord = cell.coordinate

            # Add to dictionary
            if color in color_dict:
                color_dict[color].append(coord)
            else:
                color_dict[color] = [coord]

# Print the dictionary for verification
for color, coords in color_dict.items():
    print(f"Color {color}: Coordinates {coords}")
wb = op.load_workbook("pixelart.xlsx")

pa = wb.active
print(pa.cell(1,2).value)
for i in range(1,101):
    pa.row_dimensions[i].height = 15
    col = get_column_letter(i)
    pa.column_dimensions[col].width = 2



wb.save("pixelart.xlsx")